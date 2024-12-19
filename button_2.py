from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import os

def write_to_excel(file_name, row_data):
    # Kiểm tra xem file Excel đã tồn tại chưa nếu chưa tồn tại thì tạo mới và thêm tiêu đề các cột
    if not os.path.exists(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Test Name", "Result", "Details"])
    else:  # nếu đã tồn tại thì mở ra và cho hoạt động
        workbook = load_workbook(file_name)
        sheet = workbook.active

    # Thêm dữ liệu kiểm tra ra một hàng mới
    sheet.append(row_data)
    workbook.save(file_name)
    workbook.close()

def test_double_click_me_button():
    with sync_playwright() as p:
        # Mở trình duyệt Chrome
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        # Truy cập trang
        page.goto("https://demoqa.com/buttons")

        # Nhấn đúp chuột vào nút "Double Click Me Button"
        double_click_button = page.locator("text='Double Click Me'")
        double_click_button.dblclick()

        # Kiểm tra rằng thông báo hiển thị
        try:
            assert page.inner_text("#doubleClickMessage") == "You have done a double click", "Thông báo không đúng"
            print("Thông báo hiển thị chính xác sau khi nhấn đúp chuột.")
            write_to_excel("test_results.xlsx", ["Test Double Click Me Button", "Pass", "Thông báo hiển thị chính xác sau khi nhấn đúp chuột."])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test Double Click Me Button", "Fail", str(e)])

        # Kiểm tra rằng nút không bị vô hiệu hóa
        try:
            assert double_click_button.is_enabled(), "Nút bị vô hiệu hóa"
            print("Nút có thể nhấn đúp lại và không bị vô hiệu hóa.")
            write_to_excel("test_results.xlsx", ["Test Double Click Me Button Enable", "Pass", "Nút có thể nhấn đúp lại và không bị vô hiệu hóa."])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test Double Click Me Button Enable", "Fail", str(e)])

        # Đóng trình duyệt
        browser.close()


# Gọi hàm kiểm thử
test_double_click_me_button()
