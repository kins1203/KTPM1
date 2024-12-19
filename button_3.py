from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import os

def write_to_excel(file_name, row_data):
    # Kiểm tra xem file Excel đã tồn tại chưa, nếu chưa thì tạo mới và thêm tiêu đề các cột
    if not os.path.exists(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Test Name", "Result", "Details"])
    else:  # Nếu đã tồn tại thì mở ra và cho hoạt động
        workbook = load_workbook(file_name)
        sheet = workbook.active

    # Thêm dữ liệu kiểm tra ra một hàng mới
    sheet.append(row_data)
    workbook.save(file_name)
    workbook.close()

def test_right_click_me_button():
    with sync_playwright() as p:
        # Mở trình duyệt Chrome
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        # Truy cập trang
        page.goto("https://demoqa.com/buttons")

        # Nhấn chuột phải vào nút "Right Click Me Button"
        right_click_button = page.locator("text='Right Click Me'")
        right_click_button.click(button="right")

        # Kiểm tra rằng thông báo hiển thị
        try:
            assert page.inner_text("#rightClickMessage") == "You have done a right click", "Thông báo không đúng"
            print("Thông báo hiển thị chính xác sau khi nhấn chuột phải.")
            write_to_excel("test_results.xlsx", ["Test Right Click Me Button", "Pass", "Thông báo hiển thị chính xác sau khi nhấn chuột phải."])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test Right Click Me Button", "Fail", str(e)])

        # Kiểm tra rằng nút không bị vô hiệu hóa
        try:
            assert right_click_button.is_enabled(), "Nút bị vô hiệu hóa"
            print("Nút có thể nhấn chuột phải lại và không bị vô hiệu hóa.")
            write_to_excel("test_results.xlsx", ["Test Right Click Me Button Enable", "Pass", "Nút có thể nhấn chuột phải lại và không bị vô hiệu hóa."])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test Right Click Me Button Enable", "Fail", str(e)])

        # Đóng trình duyệt
        browser.close()


# Gọi hàm kiểm thử
test_right_click_me_button()
