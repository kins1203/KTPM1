from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import os

# Hàm ghi kết quả vào file Excel
def write_to_excel(file_name, row_data):
    # Kiểm tra xem file Excel đã tồn tại chưa, nếu chưa thì tạo mới và thêm tiêu đề các cột
    if not os.path.exists(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Test Name", "Result", "Details"])
    else:  # Nếu đã tồn tại thì mở ra và cho hoạt động
        workbook = load_workbook(file_name)
        sheet = workbook.active

    # Thêm dữ liệu kiểm tra ra một hàng mới
    sheet.append(row_data)
    workbook.save(file_name)
    workbook.close()

# Hàm kiểm thử checkbox
def test_checkbox_selection():
    with sync_playwright() as p:
        # Mở trình duyệt Chrome
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        # Truy cập trang checkbox
        page.goto("https://demoqa.com/checkbox")

        # Chọn checkbox "Home"
        home_checkbox = page.locator('label[for="tree-node-home"]')
        home_checkbox.click()

        # Kiểm tra rằng checkbox "Home" có dấu tích hiển thị
        try:
            assert page.locator('#tree-node-home').is_checked(), "Checkbox Home không được chọn."
            print("Checkbox Home đã được chọn và có dấu tích hiển thị.")
            write_to_excel("test_results.xlsx", ["Test Checkbox Home", "Pass", "Checkbox Home đã được chọn và có dấu tích hiển thị."])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test Checkbox Home", "Fail", str(e)])

        # Kiểm tra hệ thống hiển thị trạng thái đã chọn
        result_text = page.locator(".display-result").inner_text()
        try:
            assert "home" in result_text.lower(), "Hệ thống không hiển thị trạng thái đã chọn."
            print("Hệ thống phản hồi chính xác với trạng thái đã chọn.")
            write_to_excel("test_results.xlsx", ["Test Result Display", "Pass", "Hệ thống phản hồi chính xác với trạng thái đã chọn."])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test Result Display", "Fail", str(e)])

        # Đóng trình duyệt
        browser.close()

# Gọi hàm kiểm thử
test_checkbox_selection()
