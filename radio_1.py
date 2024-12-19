from playwright.sync_api import sync_playwright
import time
from openpyxl import Workbook, load_workbook
import os


# Hàm ghi kết quả vào file Excel
def write_to_excel(file_name, row_data):
    # Kiểm tra xem file Excel đã tồn tại chưa, nếu chưa thì tạo mới và thêm tiêu đề các cột
    if not os.path.exists(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Test Name", "Result", "Details"])
    else:  # Nếu đã tồn tại thì mở ra và cho hoạt động
        workbook = load_workbook(file_name)
        sheet = workbook.active

    # Thêm dữ liệu kiểm tra ra một hàng mới
    sheet.append(row_data)
    workbook.save(file_name)
    workbook.close()


# Hàm kiểm thử radio button
def test_radio_button_selection():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        # Truy cập trang
        page.goto("https://demoqa.com/radio-button")

        # Bước 1: Chọn radio button Yes và kiểm tra thông báo
        page.click('label[for="yesRadio"]')
        time.sleep(1)

        try:
            # Kiểm tra thông báo đã hiển thị sau khi chọn Yes
            assert page.inner_text('#results') == "You have selected Yes", "Thông báo không đúng cho Yes"
            print("Thông báo hiển thị đúng khi chọn 'Yes'.")
            write_to_excel("test_results.xlsx", ["Test Select 'Yes'", "Pass", "Thông báo hiển thị đúng khi chọn 'Yes'."])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test Select 'Yes'", "Fail", str(e)])

        # Bước 2: Chọn radio button Impressive và kiểm tra thông báo
        page.click('label[for="impressiveRadio"]')
        time.sleep(1)

        try:
            # Kiểm tra thông báo đã hiển thị sau khi chọn Impressive
            assert page.inner_text('#results') == "You have selected Impressive", "Thông báo không đúng cho Impressive"
            print("Thông báo hiển thị đúng khi chọn 'Impressive'.")
            write_to_excel("test_results.xlsx", ["Test Select 'Impressive'", "Pass", "Thông báo hiển thị đúng khi chọn 'Impressive'."])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test Select 'Impressive'", "Fail", str(e)])

        # Bước 3: Kiểm tra rằng radio button No không thể chọn
        is_no_enabled = page.is_enabled('label[for="noRadio"]')
        try:
            assert not is_no_enabled, "Radio button No should not be selectable"
            print("Radio button 'No' không thể chọn.")
            write_to_excel("test_results.xlsx", ["Test Select 'No'", "Pass", "Radio button 'No' không thể chọn."])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test Select 'No'", "Fail", str(e)])

        # Đóng trình duyệt
        browser.close()

# Gọi hàm kiểm thử
test_radio_button_selection()
