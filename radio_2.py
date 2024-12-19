from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import os

def write_to_excel(file_name, row_data):
    # Kiểm tra xem file Excel đã tồn tại chưa nếu chưa tồn tại thì tạo mới và thêm tiêu đề các cột
    if not os.path.exists(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Test Name", "Result", "Details"])
    else:  # nêú đã tồn tại thì mở ra và cho hoạt động
        workbook = load_workbook(file_name)
        sheet = workbook.active

    # Thêm dữ liệu kiểm tra ra một hàng mới
    sheet.append(row_data)
    workbook.save(file_name)
    workbook.close()

def test_radio_buttons():
    with sync_playwright() as p:
        # Mở trình duyệt Chrome
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        # Truy cập trang
        page.goto("https://demoqa.com/radio-button")

        # Kiểm tra radio button "Yes"
        try:
            page.click('label[for="yesRadio"]')
            assert "You have selected Yes" in page.inner_text('#app'), "Thông báo không đúng cho Yes"
            print("Thông báo hiển thị đúng khi chọn 'Yes'.")
            write_to_excel("test_results.xlsx", ["Test radio button 'Yes'", "Pass", "Thông báo hiển thị đúng khi chọn 'Yes'"])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test radio button 'Yes'", "Fail", str(e)])

        # Kiểm tra radio button "Impressive"
        try:
            page.click('label[for="impressiveRadio"]')
            assert "You have selected Impressive" in page.inner_text('#app'), "Thông báo không đúng cho Impressive"
            print("Thông báo hiển thị đúng khi chọn 'Impressive'.")
            write_to_excel("test_results.xlsx", ["Test radio button 'Impressive'", "Pass", "Thông báo hiển thị đúng khi chọn 'Impressive'"])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test radio button 'Impressive'", "Fail", str(e)])

        # Kiểm tra rằng "No" không thể chọn
        try:
            no_button = page.locator('label[for="noRadio"]')
            assert no_button.is_disabled(), "Radio button 'No' không bị vô hiệu hóa"
            print("'No' không thể chọn vì bị vô hiệu hóa.")
            write_to_excel("test_results.xlsx", ["Test radio button 'No'", "Pass", "'No' không thể chọn vì bị vô hiệu hóa."])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test radio button 'No'", "Fail", str(e)])

        # Đóng trình duyệt
        browser.close()

# Gọi hàm kiểm thử
test_radio_buttons()
